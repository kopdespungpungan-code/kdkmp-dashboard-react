#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tambahkan kolom Petugas per Gondola ke template SO KDKMP (Grocery Fresh Google).
Mapping: VYRDA 1-12, PANDU 13-24, RISTA 25-36, DEDIK 37-48 (sisanya kosong)."""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = r"C:\Users\dewan\AppData\Local\hermes\profiles\coding\cache\documents\doc_6199d141b711_TEMPLATE_SO_KDKMP_UPDATE_GROCERY_EXPIRED_GOOGLE_20260803_213123.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_so_kdkmp_petugas.xlsx")

PETUGAS = [  # (start_gondola, end_gondola, nama)
    (1, 12, "VYRDA"),
    (13, 24, "PANDU"),
    (25, 36, "RISTA"),
    (37, 48, "DEDIK"),
]
def petugas_for(g):
    for a, b, n in PETUGAS:
        if a <= g <= b:
            return n
    return ""

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1E3A5F")
head_font = Font(bold=True, color="FFFFFF", size=11)
petugas_fill = PatternFill("solid", fgColor="E2F0D9")

wb = load_workbook(SRC)
if "Grocery Fresh Google" not in wb.sheetnames:
    raise SystemExit("Sheet 'Grocery Fresh Google' tidak ditemukan")
ws = wb["Grocery Fresh Google"]

# cari kolom Paraf (kolom 8 = index H) -> tambah kolom Petugas setelahnya
# header saat ini: No, Gondola, Tingkat, Nama Barang, Satuan, Qty Fisik, Keterangan, Paraf, Match Template, Status
headers = [str(c.value) for c in ws[1]]
print("Header asli:", headers)

# tambah kolom Petugas di posisi setelah Paraf (index 8 -> kolom I, tapi Match Template ada di situ)
# Lebih aman: sisipkan di kolom I (index 8, 0-based) -> pindah Match Template & Status ke kanan
ws.insert_cols(9)  # sisipkan kolom I
ws.cell(row=1, column=9, value="Petugas")
hcell = ws.cell(row=1, column=9)
hcell.fill = head_fill; hcell.font = head_font
hcell.alignment = Alignment(horizontal="center", vertical="center")
hcell.border = border

# isi petugas per baris (berdasarkan kolom Gondola = kolom 2)
for row in ws.iter_rows(min_row=2):
    gval = row[1].value  # kolom B = Gondola
    g = int(gval) if isinstance(gval, (int, float)) else None
    cell = ws.cell(row=row[0].row, column=9)
    cell.border = border
    if g is not None:
        p = petugas_for(g)
        cell.value = p if p else ""
        if p:
            cell.fill = petugas_fill
    cell.alignment = Alignment(horizontal="center")

wb.save(OUT)
print("✅ Template + Petugas tersimpan:", OUT)
print("Isi kolom Petugas: VYRDA 1-12, PANDU 13-24, RISTA 25-36, DEDIK 37-48")
