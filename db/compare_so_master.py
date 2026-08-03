#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KDKMP — BANDINGKAN SO MASTER vs OPNAME TERBARU (GROCERY FISIK).
Kolom: PRODUK | GROCERY/FISIK (TERBARU) | SISTEM (SO MASTER) | GUDANG | SELISIH | PETUGAS | GONDOLA | TANGGAL | MINGGU KE | BULAN

Sumber:
  - SO MASTER  : file SO_MASTER_UPDATED_ACUAN_JULI_2026 (sheet JULI -> STOCK AKHIR; sheet gudang -> CONVERT PCS)
  - GROCERY    : Google Sheet Grocery Fresh live (Qty Fisik = hasil opname petugas; FISIK == GROCERY, pakai yang terbaru)
  - GUDANG     : sheet 'gudang' di SO MASTER (CONVERT PCS)
Rencana user: Excel terbaru akan di-online-kan khusus pelaporan SO.

Usage:
  py -3.11 db/compare_so_master.py
Output: db/perbandingan_so_master_vs_opname.xlsx
"""
import json, os, re, urllib.request
from datetime import date, datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = r"K:/Drive Saya/KOPDES PUNGPUNGAN/GERAI/AI/laporan/Update_SO_MASTER_dan_SO_Bulanan_Juli_2026_20260801_162255/SO_MASTER_UPDATED_ACUAN_JULI_2026_20260801_162255.xlsx"
OUT = os.path.join(ROOT, "db", "perbandingan_so_master_vs_opname.xlsx")
SHEET_ID = "1V6dStO_eyyrfSw-a5686JS4SbabAAZih"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:json"

PETUGAS = [(1,12,"VYRDA"),(13,24,"PANDU"),(25,36,"RISTA"),(37,48,"DEDIK")]
def petugas_for(g):
    for a, b, n in PETUGAS:
        if a <= g <= b: return n
    return ""

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1E3A5F")
head_font = Font(bold=True, color="FFFFFF", size=11)
warn_fill = PatternFill("solid", fgColor="FDE9D9")
match_fill = PatternFill("solid", fgColor="E8F5E9")

def num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    try: return int(float(str(v).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError): return 0

def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

def get_report_info():
    """Info laporan: tanggal, minggu ke, bulan."""
    today = date.today()
    bulan_label = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"][today.month - 1]
    first_day = today.replace(day=1)
    dom = today.day
    minggu_ke = min(4, (today.day - 1) // 7 + 1)
    return {
        "tanggal": today.strftime("%d-%m-%Y"),
        "minggu_ke": min(4, (today.day - 1) // 7 + 1),
        "bulan": f"{bulan_label} {today.year}",
        "tanggal_lengkap": today.strftime("%d %B %Y")
    }

def num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    try: return int(float(str(v).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError): return 0

def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

def get_report_info():
    """Info laporan: tanggal, minggu ke, bulan."""
    today = date.today()
    bulan_label = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"][today.month - 1]
    first_day = today.replace(day=1)
    dom = today.day
    minggu_ke = min(4, (today.day - 1) // 7 + 1)
    return {
        "tanggal": today.strftime("%d-%m-%Y"),
        "minggu_ke": min(4, (today.day - 1) // 7 + 1),
        "bulan": f"{bulan_label} {today.year}",
        "tanggal_lengkap": today.strftime("%d %B %Y")
    }

PETUGAS = [(1,12,"VYRDA"),(13,24,"PANDU"),(25,36,"RISTA"),(37,48,"DEDIK")]
def petugas_for(g):
    for a, b, n in PETUGAS:
        if a <= g <= b: return n
    return ""

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1E3A5F")
head_font = Font(bold=True, color="FFFFFF", size=11)
warn_fill = PatternFill("solid", fgColor="FDE9D9")
match_fill = PatternFill("solid", fgColor="E8F5E9")

def fetch_grocery():
    """Grocery Fresh live: produk unik -> fisik (opname terbaru) + petugas + gondola."""
    print("⬇️  Mengambil Google Sheet Grocery Fresh (opname terbaru)...")
    req = urllib.request.Request(SHEET_URL, headers={"User-Agent": "Mozilla/5.0"})
    txt = urllib.request.urlopen(req, timeout=30).read().decode("utf-8", "replace")
    i, j = txt.find("{"), txt.rfind("}")
    data = json.loads(txt[i:j + 1])
    rows = data.get("table", {}).get("rows", [])
    agg = {}
    cur_gond = ""
    for r in rows:
        c = r.get("c") or []
        get = lambda k: (c[k] or {}).get("v") if k < len(c) and c[k] else None
        if get(1) is not None: cur_gond = str(get(1)).strip()
        produk = str(get(3) or "").strip()
        if not produk or produk == "-": continue
        qty = num(get(5))
        try: gond_num = int(float(cur_gond)) if cur_gond else 0
        except ValueError: gond_num = 0
        ptg = petugas_for(gond_num) if gond_num else ""
        key = norm(produk)
        if key not in agg:
            agg[key] = {"produk": produk, "fisik": 0, "petugas": ptg, "gondola": cur_gond}
        agg[key]["fisik"] += qty
        if ptg: agg[key]["petugas"] = ptg
    print(f"Produk opname: {len(agg)}")
    return agg

def load_master():
    """SO Master: sheet JULI -> produk + sistem (STOCK AKHIR/BULAN TERKHIR); sheet gudang -> gudang pcs (CONVERT PCS)."""
    if not os.path.exists(MASTER):
        print(f"⚠️ SO Master tidak ditemukan: {MASTER}")
        return [], {}
    wb = load_workbook(MASTER, data_only=True)
    master = []
    gudang_map = {}
    # sheet JULI
    if "JULI" in wb.sheetnames:
        ws = wb["JULI"]
        header = [str(c.value or "").strip() if c.value else "" for c in ws[1]]
        def hfind(*keys):
            for i, h in enumerate(header):
                hn = norm(h)
                if any(k in hn for k in keys): return i
            return None
        iNama = hfind("namabarang", "namabarang", "nama")
        iSistem = hfind("stockakhir", "stockbulanterkhir", "stock")
        iKode = hfind("kode")
        if iNama is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                nama = str(r[iNama] or "").strip() if r[iNama] else ""
                if not nama: continue
                sistem = num(r[iSistem]) if iSistem is not None and r[iSistem] is not None else 0
                kode = str(r[iKode] or "").strip() if iKode is not None and r[iKode] else ""
                master.append({"produk": nama, "sistem": sistem, "kode": kode})
            print(f"SO Master JULI: {len(master)} produk")
        # sheet gudang -> kolom CONVERT PCS (indeks 6, 0-based)
        if "gudang" in wb.sheetnames:
            ws = wb["gudang"]
            for r in ws.iter_rows(min_row=5, values_only=True):  # mulai baris 5 (data mulai)
                if not r or len(r) < 7: continue
                nama = str(r[2] or "").strip()
                if not nama: continue
                gudang_map[norm(nama)] = num(r[6])  # CONVERT PCS = indeks 6
            print(f"Gudang master: {len(gudang_map)} item, total PCS: {sum(gudang_map.values())}")
    return master, gudang_map

def match_any(n, map_norm, gf):
    """exact dulu, lalu substring (min 6 char) dari map master."""
    if n in map_norm: return map_norm[n]
    cands = []
    for k, v in map_norm.items():
        if len(k) >= 6 and (k in n or n in k):
            cands.append((abs(len(k) - len(n)), v))
    if cands:
        cands.sort(key=lambda x: x[0])
        return cands[0][1]
    return None

def main():
    gf = fetch_grocery()
    master, gudang_map = load_master()
    master_norm = {norm(m["produk"]): m for m in master}

    # Gabung: tiap produk opname terbaru (GROCERY/FISIK) + SISTEM dari master + GUDANG
    rows = []
    for n, g in gf.items():
        m = match_any(n, master_norm, gf)
        rows.append({
            "produk": g["produk"],
            "grocery": g["fisik"],          # FISIK == GROCERY (pakai yang terbaru)
            "sistem": m["sistem"] if m else 0,
            "gudang": gudang_map.get(n, 0),
            "petugas": g["petugas"],
            "gondola": g["gondola"],
        })
    # Tambahkan produk master yang belum ada di opname (biar lengkap perbandingan)
    opname_keys = {norm(r["produk"]) for r in rows}
    for m in master:
        n = norm(m["produk"])
        if n not in opname_keys:
            rows.append({
                "produk": m["produk"],
                "grocery": 0,
                "sistem": m["sistem"],
                "gudang": gudang_map.get(n, 0),
                "petugas": "",
                "gondola": "",
            })
    for r in rows:
        r["selisih"] = r["grocery"] - r["sistem"]
    rows.sort(key=lambda x: (x["selisih"] != 0, abs(x["selisih"])), reverse=True)

    # ===== Workbook =====
    wb = Workbook()
    ws = wb.active
    ws.title = "PERBANDINGAN"
    ws.sheet_view.showGridLines = False
    headers = ["PRODUK", "GROCERY/FISIK (TERBARU)", "SISTEM (SO MASTER)", "GUDANG", "SELISIH", "PETUGAS", "GONDOLA", "TANGGAL LAPORAN", "MINGGU KE", "BULAN"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    report_info = get_report_info()
    for r in rows:
        ws.append([r["produk"], r["grocery"], r["sistem"], r["gudang"], r["selisih"], r["petugas"], r["gondola"],
                   get_report_info()["tanggal"], get_report_info()["minggu_ke"], get_report_info()["bulan"]])
        row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c); cell.border = border
            if c in (2,3,4,5,7): cell.alignment = Alignment(horizontal="right")
        if r["selisih"] != 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = warn_fill
        elif r["grocery"] > 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = match_fill
    widths = [42, 22, 22, 14, 14, 14, 10, 14, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    tot_r = ws.max_row + 1
    ws.cell(row=tot_r, column=1, value="TOTAL").font = Font(bold=True)
    for col, key in [(2,"grocery"),(3,"sistem"),(4,"gudang"),(5,"selisih")]:
        total = sum(r[key] for r in rows)
        cell = ws.cell(row=tot_r, column=col, value=total)
        cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="right")
    for c in range(1, len(headers) + 1):
        ws.cell(row=tot_r, column=c).border = border

    # Sheet 2: HANYA SELISIH
    ws2 = wb.create_sheet("SELISIH SAJA")
    ws2.sheet_view.showGridLines = False
    ws2.append(["PRODUK DENGAN SELISIH (GROCERY/FISIK vs SISTEM)"])
    ws2["A1"].font = Font(bold=True, size=13, color="B45309")
    ws2.append([])
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=3, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sel = [r for r in rows if r["selisih"] != 0]
    for r in sel:
        ws2.append([r["produk"], r["grocery"], r["sistem"], r["gudang"], r["selisih"], r["petugas"], r["gondola"],
                   get_report_info()["tanggal"], get_report_info()["minggu_ke"], get_report_info()["bulan"]])
        row = ws2.max_row
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(row=row, column=c); cell.border = border; cell.fill = warn_fill
            if c in (2,3,4,5,7): cell.alignment = Alignment(horizontal="right")
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    wb.save(OUT)
    print(f"✅ Perbandingan tersimpan: {OUT}")
    print(f"Produk: {len(rows)} | Selisih != 0: {len(sel)} | Total fisik {sum(r['grocery'] for r in rows)} vs sistem {sum(r['sistem'] for r in rows)}")
    print("Sheet: PERBANDINGAN | SELISIH SAJA")

if __name__ == "__main__":
    main()