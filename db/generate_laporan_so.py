#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KDKMP — Generate LAPORAN SO (Stok Opname) format Excel.
Kolom: PRODUK | GROCERY | GUDANG | SISTEM | FISIK

Sumber utama: Google Sheet Grocery Fresh (Qty Fisik = hasil input petugas di sheet,
petugas per gondola). GROCERY/GUDANG/SISTEM dicocokkan dari sheet 'Stock Opname'
(file Excel) via fuzzy match nama produk.
Bagian khusus: DEDIK KURNIAWAN (sheet terpisah) — GUDANG GROCERY SISTEM FISIK.

Usage:
  py -3.11 db/generate_laporan_so.py
Output: db/laporan_so_kdkmp.xlsx
"""
import json, os, re, urllib.request
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = r"C:\Users\dewan\AppData\Local\hermes\profiles\coding\cache\documents\doc_4eaff82e93d9_template_so_kdkmp_petugas.xlsx"
OUT = os.path.join(ROOT, "db", "laporan_so_kdkmp.xlsx")
SHEET_ID = "1V6dStO_eyyrfSw-a5686JS4SbabAAZih"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:json"

PETUGAS = [(1,12,"VYRDA"),(13,24,"PANDU"),(25,36,"RISTA"),(37,48,"DEDIK")]
def petugas_for(g):
    for a, b, n in PETUGAS:
        if a <= g <= b: return n
    return ""

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1E3A5F")
head_font = Font(bold=True, color="FFFFFF", size=11)
fisik_fill = PatternFill("solid", fgColor="E8F5E9")
dedik_fill = PatternFill("solid", fgColor="FDE9D9")

def num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)  # float dari gviz: langsung int, jangan string-manipulasi
    try:
        return int(float(str(v).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError):
        return 0

def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

def fetch_grocery():
    """Ambil Grocery Fresh live: per produk unik -> qty fisik + petugas + gondola (gondola berlanjut)."""
    print("⬇️  Mengambil Google Sheet Grocery Fresh...")
    req = urllib.request.Request(SHEET_URL, headers={"User-Agent": "Mozilla/5.0"})
    txt = urllib.request.urlopen(req, timeout=30).read().decode("utf-8", "replace")
    i, j = txt.find("{"), txt.rfind("}")
    data = json.loads(txt[i:j + 1])
    rows = data.get("table", {}).get("rows", [])
    agg = {}
    cur_gond = ""
    for r in rows:
        c = r.get("c") or []
        get = lambda k: (c[k] or {}).get("v") if k < len(c) and c[k] else None
        if get(1) is not None:
            cur_gond = str(get(1)).strip()
        produk = str(get(3) or "").strip()
        if not produk or produk == "-":
            continue
        qty = num(get(5))
        try:
            gond_num = int(float(cur_gond)) if cur_gond else 0
        except ValueError:
            gond_num = 0
        ptg = petugas_for(gond_num) if gond_num else ""
        key = norm(produk)
        if key not in agg:
            agg[key] = {"produk": produk, "fisik": 0, "petugas": ptg, "gondola": cur_gond}
        agg[key]["fisik"] += qty
        if ptg:
            agg[key]["petugas"] = ptg
    print(f"Produk unik di Grocery Fresh: {len(agg)}")
    return agg

def load_stock():
    """Stock Opname dari Excel: produk -> {grocery, gudang, sistem}."""
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Stock Opname"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        produk = str(r[0] or "").strip()
        if not produk:
            continue
        out.append({"produk": produk, "grocery": num(r[1]), "gudang": num(r[2]), "sistem": num(r[3])})
    return out

def match_stock(n, stock_norm):
    """Cocokkan norm produk ke daftar Stock Opname: exact dulu, lalu substring."""
    if n in stock_norm:
        return stock_norm[n]
    cands = []
    for k, v in stock_norm.items():
        if len(k) >= 6 and (k in n or n in k):
            cands.append((abs(len(k) - len(n)), v))
    if cands:
        cands.sort(key=lambda x: x[0])
        return cands[0][1]
    return None

def main():
    if not os.path.exists(SRC):
        raise SystemExit(f"❌ File sumber tidak ditemukan: {SRC}")

    gf = fetch_grocery()
    stock = load_stock()
    stock_norm = {norm(s["produk"]): s for s in stock}
    print(f"Stock Opname: {len(stock)} produk")

    # Gabung: tiap produk Grocery Fresh + data stok (fuzzy)
    rows = []
    for n, g in gf.items():
        st = match_stock(n, stock_norm)
        rows.append({
            "produk": g["produk"],
            "grocery": st["grocery"] if st else 0,
            "gudang": st["gudang"] if st else 0,
            "sistem": st["sistem"] if st else 0,
            "fisik": g["fisik"],
            "petugas": g["petugas"],
            "gondola": g["gondola"],
        })
    rows.sort(key=lambda x: (x["petugas"] or "zzz", x["produk"]))

    # ===== Workbook =====
    wb = Workbook()
    ws = wb.active
    ws.title = "LAPORAN SO"
    ws.sheet_view.showGridLines = False
    headers = ["PRODUK", "GROCERY", "GUDANG", "SISTEM", "FISIK", "PETUGAS", "GONDOLA"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for s in rows:
        ws.append([s["produk"], s["grocery"], s["gudang"], s["sistem"], s["fisik"], s["petugas"], s["gondola"]])
        row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c); cell.border = border
            if c in (2,3,4,5,7): cell.alignment = Alignment(horizontal="right")
        if s["fisik"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = fisik_fill
    widths = [40, 12, 12, 12, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    tot_r = ws.max_row + 1
    ws.cell(row=tot_r, column=1, value="TOTAL").font = Font(bold=True)
    for col, key in [(2,"grocery"),(3,"gudang"),(4,"sistem"),(5,"fisik")]:
        total = sum(s[key] for s in rows)
        cell = ws.cell(row=tot_r, column=col, value=total)
        cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="right")
    for c in range(1, len(headers) + 1):
        ws.cell(row=tot_r, column=c).border = border

    # Sheet 2: DEDIK KURNIAWAN
    ws2 = wb.create_sheet("DEDIK KURNIAWAN")
    ws2.sheet_view.showGridLines = False
    ws2.append(["GUDANG GROCERY SISTEM FISIK — Khusus Diisi DEDIK KURNIAWAN"])
    ws2["A1"].font = Font(bold=True, size=13, color="B45309")
    ws2.append([])
    headers2 = ["PRODUK", "GROCERY", "GUDANG", "SISTEM", "FISIK", "GONDOLA"]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=3, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    dedik = [s for s in rows if s["petugas"].upper() == "DEDIK"]
    for s in dedik:
        ws2.append([s["produk"], s["grocery"], s["gudang"], s["sistem"], s["fisik"], s["gondola"]])
        row = ws2.max_row
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row, column=c); cell.border = border; cell.fill = dedik_fill
            if c in (2,3,4,5,6): cell.alignment = Alignment(horizontal="right")
    for i, w in enumerate([40, 12, 12, 12, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    # Sheet 3: REKAP PETUGAS
    ws3 = wb.create_sheet("REKAP PETUGAS")
    ws3.sheet_view.showGridLines = False
    ws3.append(["PETUGAS", "JUMLAH ITEM", "TOTAL FISIK", "TOTAL GROCERY", "TOTAL GUDANG", "TOTAL SISTEM"])
    for c in range(1, 7):
        cell = ws3.cell(row=1, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    agg = defaultdict(lambda: {"n":0,"fisik":0,"grocery":0,"gudang":0,"sistem":0})
    for s in rows:
        p = s["petugas"] or "—"
        agg[p]["n"] += 1
        agg[p]["fisik"] += s["fisik"]
        agg[p]["grocery"] += s["grocery"]
        agg[p]["gudang"] += s["gudang"]
        agg[p]["sistem"] += s["sistem"]
    for p, v in sorted(agg.items(), key=lambda x: -x[1]["fisik"]):
        ws3.append([p, v["n"], v["fisik"], v["grocery"], v["gudang"], v["sistem"]])
        row = ws3.max_row
        for c in range(1, 7):
            cell = ws3.cell(row=row, column=c); cell.border = border
            if c >= 2: cell.alignment = Alignment(horizontal="right")
    for i, w in enumerate([20, 12, 12, 14, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ===== Sheet 4: PER MINGGU =====
    # Rekap per petugas per Minggu 1-4 (periode laporan). Data SO adalah snapshot
    # stok opname; pembagian minggu memakai minggu berjalan periode laporan.
    from datetime import date
    today = date.today()
    bulan_label = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"][today.month - 1]
    minggu_ke = min(4, (today.day - 1) // 7 + 1)

    ws4 = wb.create_sheet("PER MINGGU")
    ws4.sheet_view.showGridLines = False
    ws4.append([f"REKAP STOK PER MINGGU — {bulan_label.upper()} {today.year}"])
    ws4["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws4.append(["Data opname snapshot; minggu berjalan: Minggu " + str(minggu_ke) + " (periode ini)."])
    ws4.append([])
    headers4 = ["MINGGU", "PETUGAS", "JUMLAH ITEM", "TOTAL FISIK", "TOTAL GROCERY", "TOTAL GUDANG", "TOTAL SISTEM"]
    ws4.append(headers4)
    for c in range(1, len(headers4) + 1):
        cell = ws4.cell(row=4, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # baris per minggu (1-4) — data ditempatkan di minggu berjalan
    for m in range(1, 5):
        for p, v in sorted(agg.items(), key=lambda x: -x[1]["fisik"]):
            ws4.append([f"Minggu {m}", p, v["n"] if m == minggu_ke else 0,
                        v["fisik"] if m == minggu_ke else 0,
                        v["grocery"] if m == minggu_ke else 0,
                        v["gudang"] if m == minggu_ke else 0,
                        v["sistem"] if m == minggu_ke else 0])
            row = ws4.max_row
            for c in range(1, len(headers4) + 1):
                cell = ws4.cell(row=row, column=c); cell.border = border
                if c >= 3: cell.alignment = Alignment(horizontal="right")
                if m == minggu_ke:
                    cell.fill = fisik_fill
    # total per minggu
    for m in range(1, 5):
        vals = [0,0,0,0,0] if m != minggu_ke else [
            sum(v["n"] for v in agg.values()),
            sum(v["fisik"] for v in agg.values()),
            sum(v["grocery"] for v in agg.values()),
            sum(v["gudang"] for v in agg.values()),
            sum(v["sistem"] for v in agg.values())]
        ws4.append([f"TOTAL Minggu {m}", "—"] + vals)
        row = ws4.max_row
        for c in range(1, len(headers4) + 1):
            cell = ws4.cell(row=row, column=c); cell.border = border
            if c >= 3: cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)
    for i, w in enumerate([16, 16, 12, 12, 14, 12, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A5"

    # ===== Sheet 5: PER BULAN =====
    ws5 = wb.create_sheet("PER BULAN")
    ws5.sheet_view.showGridLines = False
    ws5.append([f"REKAP STOK PER BULAN — {bulan_label.upper()} {today.year}"])
    ws5["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws5.append([])
    headers5 = ["BULAN", "PETUGAS", "JUMLAH ITEM", "TOTAL FISIK", "TOTAL GROCERY", "TOTAL GUDANG", "TOTAL SISTEM"]
    ws5.append(headers5)
    for c in range(1, len(headers5) + 1):
        cell = ws5.cell(row=3, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    bln = f"{bulan_label} {today.year}"
    for p, v in sorted(agg.items(), key=lambda x: -x[1]["fisik"]):
        ws5.append([bln, p, v["n"], v["fisik"], v["grocery"], v["gudang"], v["sistem"]])
        row = ws5.max_row
        for c in range(1, len(headers5) + 1):
            cell = ws5.cell(row=row, column=c); cell.border = border
            if c >= 3: cell.alignment = Alignment(horizontal="right")
    ws5.append([bln, "TOTAL",
                sum(v["n"] for v in agg.values()),
                sum(v["fisik"] for v in agg.values()),
                sum(v["grocery"] for v in agg.values()),
                sum(v["gudang"] for v in agg.values()),
                sum(v["sistem"] for v in agg.values())])
    row = ws5.max_row
    for c in range(1, len(headers5) + 1):
        cell = ws5.cell(row=row, column=c); cell.border = border
        if c >= 3: cell.alignment = Alignment(horizontal="right")
        cell.font = Font(bold=True)
    for i, w in enumerate([18, 16, 12, 12, 14, 12, 12], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = "A4"

    wb.save(OUT)
    print(f"✅ Laporan tersimpan: {OUT}")
    print(f"Produk: {len(rows)} | Fisik terisi: {sum(1 for s in rows if s['fisik'])} | DEDIK: {len(dedik)} item")
    print("Sheet: LAPORAN SO | DEDIK KURNIAWAN | REKAP PETUGAS | PER MINGGU | PER BULAN")

if __name__ == "__main__":
    main()
