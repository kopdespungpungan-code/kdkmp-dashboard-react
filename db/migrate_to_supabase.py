#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KDKMP — MIGRASI DATA ke Supabase via Management API (SQL batch + ON CONFLICT).
Sumber: SO Master Excel + Google Sheet Grocery Fresh + sheet gudang.

Butuh env:
  SUPABASE_PAT          (Personal Access Token, sbp_...)
  SUPABASE_PROJECT_REF  (default mjooqlmsswsykefaolih)

Usage:
  SUPABASE_PAT=... py -3.11 db/migrate_to_supabase.py
"""
import json, os, re, sys, uuid, hashlib, urllib.request
from datetime import date
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = r"K:/Drive Saya/KOPDES PUNGPUNGAN/GERAI/AI/laporan/Update_SO_MASTER_dan_SO_Bulanan_Juli_2026_20260801_162255/SO_MASTER_UPDATED_ACUAN_JULI_2026_20260801_162255.xlsx"
SHEET_ID = "1V6dStO_eyyrfSw-a5686JS4SbabAAZih"
REF = os.environ.get("SUPABASE_PROJECT_REF", "mjooqlmsswsykefaolih")
PAT = os.environ.get("SUPABASE_PAT", "")

PETUGAS_DEFAULT = [
    ("VYRDA", "11111111-1111-4111-8111-111111111111"),
    ("PANDU", "22222222-2222-4222-8222-222222222222"),
    ("RISTA", "33333333-3333-4333-8333-333333333333"),
    ("DEDIK", "44444444-4444-4444-8444-444444444444"),
]
GONDOLA_PETUGAS = [(1,12,"VYRDA"),(13,24,"PANDU"),(25,36,"RISTA"),(37,48,"DEDIK")]

def num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    try: return int(float(str(v).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError): return 0

def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

def run_sql(sql):
    url = f"https://api.supabase.com/v1/projects/{REF}/database/query"
    body = json.dumps({"query": sql}).encode()
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {PAT}")
    req.add_header("Content-Type", "application/json")
    req.add_header("User-Agent", "Mozilla/5.0 Chrome/126.0")
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            return resp.status, resp.read().decode("utf-8", "replace")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", "replace")

def sql_str(v):
    return "'" + str(v).replace("'", "''") + "'"

def fetch_master():
    wb = load_workbook(MASTER, data_only=True)
    products, sistem_map, gudang_map = [], {}, {}
    ws = wb["JULI"]
    header = [str(c.value or "").strip() if c.value else "" for c in ws[1]]
    def hfind(*keys):
        for i, h in enumerate(header):
            hn = norm(h)
            if any(k in hn for k in keys): return i
        return None
    iNama, iKode, iSistem = hfind("namabarang","nama"), hfind("kode"), hfind("stockakhir","stockbulanterkhir","stock")
    for r in ws.iter_rows(min_row=2, values_only=True):
        nama = str(r[iNama] or "").strip() if iNama is not None and r[iNama] else ""
        if not nama: continue
        kode = str(r[iKode] or "").strip() if iKode is not None and r[iKode] else f"UNK-{hashlib.md5(norm(nama).encode()).hexdigest()[:6].upper()}"
        sistem = num(r[iSistem]) if iSistem is not None else 0
        products.append({"kode": kode, "nama": nama})
        sistem_map[norm(nama)] = sistem
    if "gudang" in wb.sheetnames:
        ws = wb["gudang"]
        for r in ws.iter_rows(min_row=5, values_only=True):
            if not r or len(r) < 7: continue
            nama = str(r[2] or "").strip()
            if not nama: continue
            gudang_map[norm(nama)] = num(r[6])
    print(f"Master: {len(products)} produk, gudang {len(gudang_map)} item")
    return products, sistem_map, gudang_map

def fetch_opname():
    req = urllib.request.Request(f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:json", headers={"User-Agent": "Mozilla/5.0"})
    txt = urllib.request.urlopen(req, timeout=30).read().decode("utf-8", "replace")
    i, j = txt.find("{"), txt.rfind("}")
    data = json.loads(txt[i:j+1])
    rows = data.get("table", {}).get("rows", [])
    agg, cur_gond = {}, ""
    for r in rows:
        c = r.get("c") or []
        get = lambda k: (c[k] or {}).get("v") if k < len(c) and c[k] else None
        if get(1) is not None: cur_gond = str(get(1)).strip()
        produk = str(get(3) or "").strip()
        if not produk or produk == "-": continue
        qty = num(get(5))
        try: gond_num = int(float(cur_gond)) if cur_gond else 0
        except ValueError: gond_num = 0
        ptg = ""
        for a,b,n in GONDOLA_PETUGAS:
            if a <= gond_num <= b: ptg = n
        key = norm(produk)
        if key not in agg:
            agg[key] = {"nama": produk, "fisik": 0, "gondola": cur_gond, "petugas": ptg}
        agg[key]["fisik"] += qty
        if ptg: agg[key]["petugas"] = ptg
    print(f"Opname: {len(agg)} produk")
    return agg

def main():
    if not PAT:
        print("❌ SUPABASE_PAT belum diset (sbp_...)")
        sys.exit(1)

    products, sistem_map, gudang_map = fetch_master()
    opname = fetch_opname()

    # Dedupe kode (master bisa punya kode ganda) — pertahankan nama pertama
    seen, dedup = set(), []
    for p in products:
        if p["kode"] in seen: continue
        seen.add(p["kode"]); dedup.append(p)
    products = dedup

    # Gabung produk opname yang belum ada (kode OPN- deterministik dari hash nama, idempoten)
    master_norm = {norm(p["nama"]): p for p in products}
    op_extra = []
    for key, o in opname.items():
        if key not in master_norm:
            kode = "OPN-" + hashlib.md5(key.encode()).hexdigest()[:6].upper()
            op_extra.append({"kode": kode, "nama": o["nama"]})
    # Dedupe op_extra by nama
    seen_n, dedup_op = set(), []
    for p in op_extra:
        if norm(p["nama"]) in seen_n: continue
        seen_n.add(norm(p["nama"])); dedup_op.append(p)
    op_extra = dedup_op
    all_products = products + op_extra
    print(f"Total produk gabungan (dedupe): {len(all_products)} (master {len(products)} + opname baru {len(op_extra)})")

    # ===== 1. PRODUCTS (INSERT ON CONFLICT DO NOTHING) =====
    vals = ",\n".join(f"({sql_str(p['kode'])}, {sql_str(p['nama'])}, 'PCS')" for p in all_products)
    sql = f"INSERT INTO products (kode, nama, satuan) VALUES {vals} ON CONFLICT (kode) DO NOTHING;"
    st, body = run_sql(sql)
    print(f"products: HTTP {st} {body[:120]}")
    if st != 201:
        print("⚠️ Gagal insert products, lihat error."); return

    # Ambil id_by_norm dari DB (semua produk termasuk yang sudah ada)
    st, body = run_sql("SELECT id, kode, nama FROM products")
    prods = json.loads(body)
    id_by_norm = {norm(p["nama"]): p["id"] for p in prods}
    id_by_kode = {p["kode"]: p["id"] for p in prods}
    print(f"Products di DB: {len(prods)}")

    # ===== 2. STOCK_SYSTEM =====
    rows = []
    for key, sistem in sistem_map.items():
        pid = id_by_norm.get(key)
        if pid: rows.append(f"({sql_str(pid)}, {sistem})")
    if rows:
        vals = ",\n".join(rows)
        sql = f"INSERT INTO stock_system (produk_id, stok_akhir) VALUES {vals} ON CONFLICT (produk_id) DO UPDATE SET stok_akhir = EXCLUDED.stok_akhir;"
        st, body = run_sql(sql)
        print(f"stock_system: HTTP {st} {body[:120]}")

    # ===== 3. GUDANG (fuzzy match nama sheet gudang -> produk) =====
    id_norm_list = {norm(p["nama"]): p["id"] for p in prods}
    rows = []
    for key, qty in gudang_map.items():
        pid = id_norm_list.get(key)
        if not pid:  # fuzzy: substring min 6 char
            cands = []
            for k, v in id_norm_list.items():
                if len(k) >= 6 and (k in key or key in k):
                    cands.append((abs(len(k) - len(key)), v))
            if cands:
                cands.sort(key=lambda x: x[0])
                pid = cands[0][1]
        if pid:
            rows.append((pid, qty))
    if rows:
        vals = ",\n".join(f"({sql_str(pid)}, {qty}, CURRENT_DATE)" for pid, qty in rows)
        sql = f"INSERT INTO gudang (produk_id, qty_pcs, tanggal_rekap) VALUES {vals} ON CONFLICT (produk_id) DO UPDATE SET qty_pcs = EXCLUDED.qty_pcs;"
        st, body = run_sql(sql)
        print(f"gudang: HTTP {st} {body[:120]} (matched {len(rows)}/{len(gudang_map)})")

    # ===== 3b. Pastikan constraint UNIQUE(produk_id, tanggal) untuk ON CONFLICT =====
    st, body = run_sql("""DO $$ BEGIN
        IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'stock_opname_produk_tanggal_key') THEN
            ALTER TABLE stock_opname ADD CONSTRAINT stock_opname_produk_tanggal_key UNIQUE (produk_id, tanggal);
        END IF;
    END $$;""")
    print(f"constraint stock_opname: HTTP {st} {body[:120]}")

    # ===== 4. STOCK_OPNAME (tanggal hari ini) =====
    today = date.today().isoformat()
    minggu_ke = min(4, (date.today().day - 1) // 7 + 1)
    bulan = date.today().strftime("%B %Y")
    rows = []
    for key, o in opname.items():
        pid = id_by_norm.get(key)
        if not pid: continue
        petugas_id = None
        for _, _, n in GONDOLA_PETUGAS:
            if o["petugas"] == n:
                petugas_id = dict(PETUGAS_DEFAULT)[n]
        rows.append(f"({sql_str(pid)}, {sql_str(petugas_id) if petugas_id else 'NULL'}, {num(o['gondola'])}, {o['fisik']}, {sql_str(today)}, {minggu_ke}, {sql_str(bulan)})")
    if rows:
        vals = ",\n".join(rows)
        sql = f"""INSERT INTO stock_opname (produk_id, petugas_id, gondola, qty_fisik, tanggal, minggu_ke, bulan)
VALUES {vals}
ON CONFLICT (produk_id, tanggal) DO UPDATE SET qty_fisik = EXCLUDED.qty_fisik, petugas_id = EXCLUDED.petugas_id, gondola = EXCLUDED.gondola;"""
        st, body = run_sql(sql)
        print(f"stock_opname: HTTP {st} {body[:120]}")

    # ===== 5. PETUGAS seed =====
    vals = ",\n".join(f"({sql_str(uid)}, {sql_str(nama)}, 'seed')" for nama, uid in PETUGAS_DEFAULT)
    sql = f"INSERT INTO petugas (id, nama, pin_hash) VALUES {vals} ON CONFLICT (nama) DO NOTHING;"
    st, body = run_sql(sql)
    print(f"petugas: HTTP {st} {body[:120]}")

    # ===== Rekap =====
    for q in ["SELECT count(*) FROM products", "SELECT count(*) FROM stock_system", "SELECT count(*) FROM gudang", "SELECT count(*) FROM stock_opname", "SELECT count(*) FROM petugas"]:
        st, body = run_sql(q)
        print(f"{q}: {body}")

    print("\n✅ MIGRASI SELESAI!")

if __name__ == "__main__":
    main()