#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Buat template Excel SO (Stok Opname) + SO/APN untuk KDKMP."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_so_kdkmp.xlsx")

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1E3A5F")
head_font = Font(bold=True, color="FFFFFF", size=11)
ex_fill = PatternFill("solid", fgColor="FFF7E6")

wb = Workbook()

# ===== Sheet 1: Stock Opname =====
ws = wb.active
ws.title = "Stock Opname"
ws.sheet_view.showGridLines = False

ws.append(["Produk", "Stock Grocery", "Stock Gudang", "Stock On System", "Expired", "Keterangan"])
ws.append(["Beras 5kg", 12, 8, 19, "31/12/2026", "Contoh: fisik 20, sistem 19 -> selisih +1"])
ws.append(["Minyak Goreng 1L", 30, 15, 46, "15/09/2026", ""])
ws.append(["Gula 1kg", 25, 10, 35, "31/12/2027", ""])
ws.append(["Telur 1kg", 0, 6, 6, "05/08/2026", "Expired dekat — periksa"])
ws.append(["Sabun Mandi", 40, 20, 60, "", ""])
ws.append(["", "", "", "", "", "Isi sesuai stok opname kamu"])

# header style
for c in range(1, 7):
    cell = ws.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
# contoh rows
for r in range(2, 8):
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        if r <= 6: cell.fill = ex_fill
        if c in (2, 3, 4, 5): cell.alignment = Alignment(horizontal="right")
# lebar kolom
widths = [22, 16, 16, 18, 16, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ===== Sheet 2: SO / APN (induk) =====
ws2 = wb.create_sheet("SO APN")
ws2.sheet_view.showGridLines = False
ws2.append(["Nomor SO", "Tanggal", "Customer", "Keterangan", "Total", "Status"])
ws2.append(["SO-2026-001", "01/08/2026", "Toko Sumber Rejeki", "Pesanan awal", 250000, "open"])
ws2.append(["SO-2026-002", "02/08/2026", "Warung Bu Tini", "", 175000, "selesai"])
for c in range(1, 7):
    cell = ws2.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
for r in range(2, 4):
    for c in range(1, 7):
        cell = ws2.cell(row=r, column=c); cell.border = border; cell.fill = ex_fill
for i, w in enumerate([14, 14, 22, 20, 14, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ===== Sheet 3: Item SO / APN =====
ws3 = wb.create_sheet("Item SO APN")
ws3.sheet_view.showGridLines = False
ws3.append(["Nomor SO", "Produk", "Qty", "Harga Satuan", "Subtotal"])
ws3.append(["SO-2026-001", "Beras 5kg", 10, 15500, 155000])
ws3.append(["SO-2026-001", "Minyak Goreng 1L", 5, 19000, 95000])
ws3.append(["SO-2026-002", "Gula 1kg", 10, 17500, 175000])
for c in range(1, 6):
    cell = ws3.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
for r in range(2, 5):
    for c in range(1, 6):
        cell = ws3.cell(row=r, column=c); cell.border = border; cell.fill = ex_fill
for i, w in enumerate([14, 18, 10, 16, 14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ===== Sheet 4: Petunjuk =====
ws4 = wb.create_sheet("Petunjuk")
ws4.sheet_view.showGridLines = False
ws4.append(["PANDUAN CEPAT TEMPLATE SO KDKMP"])
ws4.append([])
ws4.append(["1. Sheet 'Stock Opname' — untuk data stok opname (Grocery/Gudang/On System)."])
ws4.append(["   Dashboard mengenali nama kolom otomatis (boleh diganti, asal mengandung kata):"])
ws4.append(["      Produk       -> 'produk' / 'nama' / 'barang' / 'item'"])
ws4.append(["      Stock Grocery-> 'grocery' / 'etalase' / 'toko'"])
ws4.append(["      Stock Gudang -> 'gudang' / 'warehouse'"])
ws4.append(["      On System    -> 'on system' / 'sistem' / 'system' / 'on'"])
ws4.append(["      Expired      -> 'expired' / 'kadaluarsa' / 'kedaluwarsa' / 'tanggal kadaluarsa' (format dd/mm/yyyy)"])
ws4.append([])
ws4.append(["2. Sheet 'SO APN' — data sales order / APN induk."])
ws4.append(["3. Sheet 'Item SO APN' — detail produk per SO."])
ws4.append([])
ws4.append(["4. CARA PAKAI DENGAN GOOGLE FORM:"])
ws4.append(["   a. Buka sheets.new, lalu File > Import > Upload file ini."])
ws4.append(["   b. Buat Google Form > Responses > pilih sheet ini sebagai tujuan."])
ws4.append(["   c. Kirim link/ID sheet ke tim KDKMP -> dipasang di src/config.js (SO_SHEET_ID)."])
ws4.append([])
ws4.append(["5. Selisih otomatis = (Grocery + Gudang) - On System. Baris merah = ada selisih."])
ws4.append(["   File dibuat oleh: SNP (Sumber Niaga Prima) / Dedik Kurniawan — Dashboard KDKMP."])
for i, w in enumerate([8, 110], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4["A1"].font = Font(bold=True, size=13, color="1E3A5F")

wb.save(OUT)
print("✅ Template tersimpan:", OUT)
