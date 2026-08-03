#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Buat template SO KDKMP terisi data dari Google Sheet (gondola/rak/produk/qty/expired)
+ kolom Petugas (per gondola). Sumber JSON: so_sheet.json (hasil fetch gviz)."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
JSON = os.path.join(ROOT, "so_sheet.json")
OUT = os.path.join(ROOT, "template_so_kdkmp_isi.xlsx")

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1E3A5F")
head_font = Font(bold=True, color="FFFFFF", size=11)
gond_fill = PatternFill("solid", fgColor="E8F0FA")
empty_fill = PatternFill("solid", fgColor="F5F5F5")

def load_rows():
    txt = open(JSON, encoding="utf-8").read()
    i, j = txt.find("{"), txt.rfind("}")
    data = json.loads(txt[i:j + 1])
    rows = data.get("table", {}).get("rows", [])
    out = []
    cur_gond = None
    for r in rows:
        c = r.get("c") or []
        get = lambda k: (c[k] or {}).get("v") if k < len(c) and c[k] else None
        g = get(1)
        if g is not None:
            cur_gond = int(g)
        rak = get(2)
        produk = str(get(3) or "").strip()
        satuan = str(get(4) or "").strip()
        qty = get(5)
        exp = str(get(6) or "").strip()
        out.append({
            "gondola": cur_gond if cur_gond is not None else "",
            "rak": int(rak) if rak is not None else "",
            "produk": "" if produk == "-" else produk,
            "satuan": "" if satuan == "-" else satuan,
            "qty": qty,
            "expired": "" if exp == "-" else exp,
            "petugas": "",  # KOLOM BARU: diisi per gondola
        })
    return out

def main():
    data = load_rows()
    print(f"Baris dibaca: {len(data)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Opname"
    ws.sheet_view.showGridLines = False

    headers = ["Gondola", "Rak", "Produk", "Satuan", "Qty", "Expired", "Petugas"]
    ws.append(headers)
    for hc in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=hc)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # blok gondola: beri warna header gondola + petugas kosong
    for r, row in enumerate(data, start=2):
        ws.append([
            row["gondola"] if row["rak"] == 1 else "",
            row["rak"], row["produk"], row["satuan"],
            row["qty"], row["expired"], row["petugas"],
        ])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c in (5, 6, 7):
                cell.alignment = Alignment(horizontal="right" if c == 5 else "left")
        # baris pertama tiap gondola diberi warna latar
        if row["rak"] == 1:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = gond_fill

    widths = [10, 6, 30, 8, 8, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Sheet petunjuk
    ws2 = wb.create_sheet("Petunjuk")
    ws2.sheet_view.showGridLines = False
    lines = [
        "PANDUAN TEMPLATE SO KDKMP (terisi dari Google Sheet)",
        "",
        "1. Sheet 'Stock Opname' berisi data gondola/rak/produk asli (60 gondola x 6 rak).",
        "2. KOLOM BARU 'Petugas' — isi nama petugas yang bertanggung jawab PER GONDOLA.",
        "   (isi pada baris pertama tiap blok gondola yang berwarna biru muda, atau semua baris gondola tsb)",
        "3. 'Qty' = jumlah stok. 'Expired' = tanggal kadaluarsa (format exp dd-mm-yy / dd-mm-yyyy).",
        "4. Baris 'Produk' kosong / '-' = rak kosong.",
        "5. Dashboard mendeteksi kolom otomatis: gondola, rak, produk, qty, expired, petugas.",
        "",
        "Dibuat oleh: SNP (Sumber Niaga Prima) / Dedik Kurniawan — Dashboard KDKMP.",
    ]
    for ln in lines:
        ws2.append([ln])
    ws2["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws2.column_dimensions["A"].width = 100

    wb.save(OUT)
    print("✅ Template terisi tersimpan:", OUT)

if __name__ == "__main__":
    main()
